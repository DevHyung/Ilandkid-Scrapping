#-*- encoding:utf8 -*-
import os
from selenium import webdriver
from bs4 import BeautifulSoup
from urllib.request import urlretrieve
from openpyxl import load_workbook, Workbook

def Login(): # 로그인하는 함수
    global driver

    url = './chromedriver'  # 드라이브가 있는 경로
    driver = webdriver.Chrome(url)
    driver.get("http://ilandkid.com/login2.php")

    driver.find_element_by_xpath("//input[@name='am_id']").send_keys("YOUERID")  # ID 입력
    driver.find_element_by_xpath("//input[@name='am_pwd']").send_keys("YOUEPW")  # PW 입력
    driver.execute_script("javascript:go_login();return false;")

def getdata(FName, route, rowNum):  # 페이지로 들어가서 정보를 갖고 오는 함수 & 그 정보를 엑셀에 저장
    # 상품명
    try:
        elem1 = driver.find_element_by_xpath("/html/body/form/div/div/div/div[1]/div[1]/div[1]/div[2]/font/b").text
    except :
        elem1 = driver.find_element_by_xpath("/html/body/form/div/div[1]/div/div/div[1]/div[1]/div[2]/font/b").text

    # 공급사 상품명
    elem2 = driver.find_element_by_xpath("/html/body/form/div/div[1]/div/div/div[1]/div[2]/table/tbody/tr/td[3]/div[1]/table/tbody/tr[1]/td[2]").text

    # 상품명(관리용)
    elem3 = driver.find_element_by_xpath("/html/body/form/div/div[1]/div/div/div[1]/div[2]/table/tbody/tr/td[3]/div[1]/table/tbody/tr[2]/td[2]").text

    # 공급가
    elem4 = driver.find_element_by_xpath("/html/body/form/div/div[1]/div/div/div[1]/div[2]/table/tbody/tr/td[3]/div[1]/table/tbody/tr[5]/td[2]").text
    elem4 = elem4.replace(" 원 (부가세 별도)", "")

    # 등록일
    elem5 = driver.find_element_by_xpath("/html/body/form/div/div[1]/div/div/div[1]/div[2]/table/tbody/tr/td[3]/div[1]/table/tbody/tr[8]/td[2]").text

    # 옵션 입력 - 색상
    bs4 = BeautifulSoup(driver.page_source, "html.parser")
    List = bs4.find('td', style='PADDING-LEFT:4px;').find('select').findAll('option')

    str6 = ''
    for i in range(0, len(List)):
        str6 = str6 + List[i]['value'] + "|"
    str6 = str6[:len(str6)-1]

    # 옵션 입력 - 사이즈 (치수)
    try:
        elem7 = driver.find_element_by_xpath("/html/body/form/div/div[1]/div/div/div[1]/div[2]/table/tbody/tr/td[3]/table/tbody/tr[2]/td/table/tbody/tr[3]/td[2]/table/tbody/tr").text
        elem7 = elem7.replace('\n', '|').replace('-', '|')
    except:
        elem7 = ''

    # 옵션 입력 - 사이즈 (사이즈)
    if elem7 == '':
        try:
            elem7 = driver.find_element_by_xpath("/html/body/form/div/div[1]/div/div/div[1]/div[2]/table/tbody/tr/td[3]/div[1]/table/tbody/tr[7]/td[2]").text
            elem7 = elem7.replace('호', '').replace('-', '|').replace(',', '|').replace(' ', '')
        except:
            elem7 = ''

    try:
        soldOut = driver.find_element_by_xpath('/html/body/form/div/div/div/div[1]/div[1]/div[1]/div[2]/font[2]').text
        if soldOut != '[품절]': soldOut = ''
    except:
        soldOut = ""

    ws = wb.active
    ws.cell(row=rowNum, column=1, value=route) # 공급사(경로)
    ws.cell(row=rowNum, column=2, value=elem1) # 상품명
    ws.cell(row=rowNum, column=3, value=elem3) # 상품명(관리용)
    ws.cell(row=rowNum, column=4, value=elem2) # 공급사 상품명
    ws.cell(row=rowNum, column=5, value=elem4) # 공급가
    ws.cell(row=rowNum, column=8, value="색상{" + str6 + "}//사이즈{" + elem7 + "}") # 옵션입력
    ws.cell(row=rowNum, column=9, value=elem5) # 등록일
    ws.cell(row=rowNum, column=10, value=soldOut) # 품절
    wb.save(FName + '/' + '#' + FName + '.xlsx')

def getimg(FName, imgNum) : # 이미지 저장하는 함수
    try:
        bs4 = BeautifulSoup(driver.page_source, "html.parser")
        img1 = bs4.find('div', id='pic1').find('img')
        img1 = "http://ilandkid.com/" + img1['src'][2:]
        urlretrieve(img1, FName + "/" + str(imgNum).rjust(5, '0')+ "_img01.jpg")

        img2 = bs4.find('div', align='center').findAll('img')
        for i in range(0, len(img2)):
            rink = "http://ilandkid.com/" + img2[i]['src'][2:]
            urlretrieve(rink, FName + "/" + str(imgNum).rjust(5, '0')+ "_img" + str(i + 2).rjust(2, '0') + ".jpg")
    except:
        bs4 = ''

if __name__=="__main__":
    Title = ['부르뎅', '마마아동복', '포키아동복', '탑랜드', '서울원아동복', '크레용아동복', '페인트타운', '초특가이월', '주니어브랜드']

    Login()
    # title_01 = '탑랜드'
    # title_02 = '홈런'
    # title_03 = '홈런(17가을1차)'
    # title_04 = '원피스/치마 (Dress-Skirt)'
    for i in range(0, len(Title)): # 최상위 카테고리 (부르뎅)
        #if i != 0: Login()
        global wb
        rowNum = 2
        imgNum = 2

        temproute1 = '"' + Title[i] + '"'
        while True:
            answer = input(temproute1 + "의 데이터를 추출하시겠습니까? y/n : ")
            if answer == "y":

                driver.execute_script("javascript:OnDisplayToggle('idMenu" + str(i) + "');")
                # if Title[i] != title_01: continue

                if not os.path.isdir(Title[i]):
                    os.mkdir(Title[i])

                wb = Workbook()
                ws = wb.active

                ws.cell(row=1, column=1, value='공급사(경로)')  # 공급사(경로)
                ws.cell(row=1, column=2, value='상품명')  # 상품명
                ws.cell(row=1, column=3, value='상품명(관리용)')  # 상품명(관리용)
                ws.cell(row=1, column=4, value='공급사 상품명')  # 공급사 상품명
                ws.cell(row=1, column=5, value='공급가')  # 공급가
                ws.cell(row=1, column=6, value='변동가')  # 변동가
                ws.cell(row=1, column=7, value='공급가변환($->￦)')  # 공급가변환($->￦)
                ws.cell(row=1, column=8, value='옵션입력')  # 옵션입력
                ws.cell(row=1, column=9, value='등록일')  # 등록일
                ws.cell(row=1, column=10, value='품절')  # 품절
                ws.cell(row=1, column=11, value='기타')  # 기타

                wb.save(Title[i] + '/' + '#' + Title[i] + '.xlsx')
                break
            elif answer == "n":
                break
            else:
                continue

        if answer == 'y' :
            bs4 = BeautifulSoup(driver.page_source, "html.parser")
            List1 = bs4.find('div', id='idMenu' + str(i)).find('table').find('table').findAll('td')

            for j in range(0, len(List1)): # 1차 하위 카테고리 (AN)
                if j != 0:
                    print("\n--- 페이지를 로딩중입니다. 잠시만 기다려주십시오.\n")
                    driver.close()
                    Login()
                    driver.execute_script("javascript:OnDisplayToggle('idMenu"+ str(i) +"');")

                data = driver.find_elements_by_xpath("//*[@id='idMenu" + str(i) + "']/table/tbody/tr/td/table/tbody/tr/td/a/font")

                # try:
                data_01 = data[j].text.replace('\n', '').replace('  ', ' ').strip()
                # print(data_01)
                # if data_01 != title_02: continue
                temproute2 = '"' + Title[i] + '"-"' + data_01 + '"'
                while True:
                    answer = input(temproute2 + "의 데이터를 추출하시겠습니까? y/n : ")
                    if answer == "y":
                        data[j].click()
                        break
                    elif answer == "n":
                        break
                    else:
                        continue
                # except:
                #     continue

                if answer == 'y':
                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                    List2 = bs4.findAll('td', height='21')

                    for k in range(0, len(List2)): # 2차 하위 카테고리 (17여름세일)
                        data = driver.find_elements_by_xpath("//table[3]/tbody/tr/td[2]/table/tbody/tr/td/a/font")
                        data_02 = data[k].text.replace('\n', '').replace('  ', ' ').strip()
                        # if data_02 != title_03: continue
                        temproute3 = '"' + Title[i] + '"-"' + data_01 + '"-"' + data_02 + '"'
                        while True:
                            answer = input(temproute3 + "의 데이터를 추출하시겠습니까? y/n : ")
                            if answer == "y":
                                data[k].click()
                                break
                            elif answer == "n":
                                break
                            else:
                                continue

                        if answer == 'y':
                            bs4 = BeautifulSoup(driver.page_source, "html.parser")
                            List3 = bs4.findAll('td', style='padding-right:60px')

                            for l in range(0, len(List3)): # 3차 하위 카테고리 (원피스)
                                data = driver.find_elements_by_xpath("//table[4]/tbody/tr/td/table/tbody/tr/td/a")
                                data_03 = data[l].text.replace('\n', '').replace('  ', ' ').strip()
                                # if data_03 != title_04: continue
                                route = Title[i] + "-" + data_01 + '-' + data_02 + '-' + data_03
                                data[l].click()

                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                List4 = bs4.findAll('td', width='200')
                                paging = bs4.findAll('table', {'class': 'paging'})

                                row = int(len(List4) / 5) + 1
                                endrow = len(List4) % 5

                                if (len(paging) == 2):  # 페이지 넘김 있음
                                    while True:
                                        if (endrow != 0):  # 항목이 5개로 나눠떨어지지 않음
                                            for p in range(0, row):
                                                if (p != row - 1):
                                                    for o in range(0, 5):
                                                        driver.find_element_by_xpath(
                                                            "/html/body/form/div/div/div/div[1]/table[8]/tbody/tr[" + str(
                                                                p * 2 + 1) + "]/td[" + str(
                                                                o * 2 + 1) + "]/table/tbody/tr[1]/td/a/img").click()
                                                        getdata(Title[i], route, rowNum)
                                                        getimg(Title[i], imgNum)
                                                        rowNum = rowNum + 1
                                                        imgNum = imgNum + 1
                                                        driver.back()
                                                else:
                                                    for o in range(0, len(List4) % 5):
                                                        driver.find_element_by_xpath(
                                                            "/html/body/form/div/div/div/div[1]/table[8]/tbody/tr[" + str(
                                                                p * 2 + 1) + "]/td[" + str(
                                                                o * 2 + 1) + "]/table/tbody/tr[1]/td/a/img").click()
                                                        getdata(Title[i], route, rowNum)
                                                        getimg(Title[i], imgNum)
                                                        rowNum = rowNum + 1
                                                        imgNum = imgNum + 1
                                                        driver.back()
                                            try:
                                                driver.find_element_by_xpath(
                                                    "/html/body/form/div/div/div/div[1]/table[7]/tbody/tr/td[4]/a").click()  # 다음으로 넘어가기
                                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                List5 = bs4.findAll('td', width='200')
                                                paging = bs4.findAll('table', {'class': 'paging'})

                                                row = int(len(List5) / 5) + 1
                                                endrow = len(List5) % 5
                                            except:
                                                break

                                        else:  # 항목이 5개로 나눠떨어짐
                                            for p in range(0, row - 1):
                                                for o in range(0, 5):
                                                    driver.find_element_by_xpath(
                                                        "/html/body/form/div/div/div/div[1]/table[8]/tbody/tr[" + str(
                                                            p * 2 + 1) + "]/td[" + str(
                                                            o * 2 + 1) + "]/table/tbody/tr[1]/td/a/img").click()
                                                    getdata(Title[i], route, rowNum)
                                                    getimg(Title[i], imgNum)
                                                    rowNum = rowNum + 1
                                                    imgNum = imgNum + 1
                                                    driver.back()
                                            try:
                                                driver.find_element_by_xpath(
                                                    "/html/body/form/div/div/div/div[1]/table[6]/tbody/tr/td[4]/a").click()  # 다음으로 넘어가기
                                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                List6 = bs4.findAll('td', width='200')
                                                paging = bs4.findAll('table', {'class': 'paging'})

                                                row = int(len(List6) / 5) + 1
                                                endrow = len(List6) % 5
                                            except:
                                                break

                                else:  # 페이지 넘김 없음
                                    if (endrow != 0):  # 항목이 5개로 나눠떨어지지 않음
                                        for p in range(0, row):
                                            if (p != row - 1):
                                                for o in range(0, 5):
                                                    driver.find_element_by_xpath(
                                                        "/html/body/form/div/div/div/div[1]/table[7]/tbody/tr[" + str(
                                                            p * 2 + 1) + "]/td[" + str(
                                                            o * 2 + 1) + "]/table/tbody/tr[1]/td/a/img").click()
                                                    getdata(Title[i], route, rowNum)
                                                    getimg(Title[i], imgNum)
                                                    rowNum = rowNum + 1
                                                    imgNum = imgNum + 1
                                                    driver.back()
                                            else:
                                                for o in range(0, len(List4) % 5):
                                                    driver.find_element_by_xpath(
                                                        "/html/body/form/div/div/div/div[1]/table[7]/tbody/tr[" + str(
                                                            p * 2 + 1) + "]/td[" + str(
                                                            o * 2 + 1) + "]/table/tbody/tr[1]/td/a/img").click()
                                                    getdata(Title[i], route, rowNum)
                                                    getimg(Title[i], imgNum)
                                                    rowNum = rowNum + 1
                                                    imgNum = imgNum + 1
                                                    driver.back()

                                    else:  # 항목이 5개로 나눠떨어짐
                                        for p in range(0, row - 1):
                                            for o in range(0, 5):
                                                driver.find_element_by_xpath(
                                                    "/html/body/form/div/div/div/div[1]/table[7]/tbody/tr[" + str(
                                                        p * 2 + 1) + "]/td[" + str(
                                                        o * 2 + 1) + "]/table/tbody/tr[1]/td/a/img").click()
                                                getdata(Title[i], route, rowNum)
                                                getimg(Title[i], imgNum)
                                                rowNum = rowNum + 1
                                                imgNum = imgNum + 1
                                                driver.back()
                        elif answer == 'n':
                            continue

                elif answer == 'n':
                    continue

        elif answer == 'n' :
            continue

        try: wb.close()
        except: elem = ''